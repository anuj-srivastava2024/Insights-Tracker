import streamlit as st
import json
import os
from datetime import date, datetime
import pandas as pd
from io import BytesIO
import anthropic

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Trade Contract Tracker",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

DATA_FILE = "data.json"

TAGS = {
    "Absorption":     "#6366f1",
    "Iceberg":        "#0ea5e9",
    "Tool":           "#f59e0b",
    "Extreme Delta":  "#ef4444",
    "Extreme Volume": "#8b5cf6",
    "Bullish":        "#22c55e",
    "Bearish":        "#f97316",
}

# ── Persistence ───────────────────────────────────────────────────────────────
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {}

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)

def export_to_excel(db):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for page_name, page_data in db.items():
            columns = page_data.get("columns", [])
            rows = page_data.get("rows", {})
            if not columns and not rows:
                continue
            records = []
            for dstr in sorted(rows.keys()):
                try:
                    d = datetime.strptime(dstr, "%Y-%m-%d")
                    date_display = d.strftime("%a, %b %d %Y")
                except:
                    date_display = dstr
                row_record = {"Date": date_display}
                for col in columns:
                    cell = rows[dstr].get(col, {"text": "", "tags": []})
                    text = cell.get("text", "")
                    tags = ", ".join(cell.get("tags", []))
                    row_record[col] = f"{text}\n[{tags}]" if tags else text
                records.append(row_record)
            if not records:
                records = [{"Date": "No data yet"}]
            df = pd.DataFrame(records)
            sheet_name = page_name[:31].replace("/", "-").replace("\\", "-").replace("*","").replace("?","").replace("[","").replace("]","").replace(":","")
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1E293B")
            header_font = Font(color="94A3B8", bold=True, size=10)
            date_fill   = PatternFill("solid", fgColor="1E293B")
            date_font   = Font(color="6366F1", bold=True, size=10)
            cell_font   = Font(color="E2E8F0", size=10)
            thin_border = Border(
                left=Side(style="thin", color="334155"), right=Side(style="thin", color="334155"),
                top=Side(style="thin", color="334155"),  bottom=Side(style="thin", color="334155"),
            )
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = 12 if col_idx == 1 else 30
            for row in ws.iter_rows(min_row=2):
                for col_idx, cell in enumerate(row, 1):
                    cell.font = date_font if col_idx == 1 else cell_font
                    cell.fill = date_fill if col_idx == 1 else PatternFill("solid", fgColor="0F172A")
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[row[0].row].height = 60
    output.seek(0)
    return output

# ── Build tracker context string for Claude ───────────────────────────────────
def build_tracker_context(db, focus_page=None):
    """Serialise all tracker data (or one page) into a readable text block."""
    lines = ["=== TRADE TRACKER DATA ===\n"]
    pages = {focus_page: db[focus_page]} if focus_page and focus_page in db else db
    for page_name, page_data in pages.items():
        lines.append(f"\n## Page: {page_name}")
        columns = page_data.get("columns", [])
        rows    = page_data.get("rows", {})
        if not columns:
            lines.append("  (no columns defined)")
            continue
        lines.append(f"  Contracts tracked: {', '.join(columns)}")
        if not rows:
            lines.append("  (no data rows yet)")
            continue
        for dstr in sorted(rows.keys()):
            try:
                d = datetime.strptime(dstr, "%Y-%m-%d")
                date_label = d.strftime("%a %b %d %Y")
            except:
                date_label = dstr
            lines.append(f"\n  📅 {date_label}")
            row_data = rows[dstr]
            for col in columns:
                cell = row_data.get(col, {})
                text = cell.get("text", "").strip()
                tags = cell.get("tags", [])
                if text or tags:
                    tag_str = f"  [Tags: {', '.join(tags)}]" if tags else ""
                    lines.append(f"    • {col}: {text}{tag_str}")
                else:
                    lines.append(f"    • {col}: (empty)")
    lines.append("\n=== END OF TRACKER DATA ===")
    return "\n".join(lines)

# ── Call Claude API ───────────────────────────────────────────────────────────
def ask_claude(messages, tracker_context):
    api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return "⚠️ No Anthropic API key found. Add `ANTHROPIC_API_KEY` to your Streamlit secrets."
    try:
        client = anthropic.Anthropic(api_key=api_key)
        system_prompt = f"""You are an expert commodity and futures trade analyst assistant.
You have access to the user's trade observation tracker data below.
Each page represents a trading session log. Each column is a contract/product.
Observations include free-text notes and tags like: Absorption, Iceberg, Tool, Extreme Delta, Extreme Volume, Bullish, Bearish.

Use this data to answer questions, identify patterns, and provide actionable trade suggestions.
Be specific — reference actual dates, contracts, and observations from the data.
When suggesting trades, explain your reasoning based on the patterns you see.

{tracker_context}"""

        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=1500,
            system=system_prompt,
            messages=messages,
        )
        return response.content[0].text
    except Exception as e:
        return f"⚠️ Claude API error: {e}"

# ── Session state ─────────────────────────────────────────────────────────────
if "db" not in st.session_state:
    st.session_state.db = load_data()
if "current_page" not in st.session_state:
    st.session_state.current_page = None
if "active_cell" not in st.session_state:
    st.session_state.active_cell = None
if "edits" not in st.session_state:
    st.session_state.edits = {}
if "has_unsaved" not in st.session_state:
    st.session_state.has_unsaved = False
if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = []   # [{role, content}]
if "chat_focus" not in st.session_state:
    st.session_state.chat_focus = "All Pages"

db = st.session_state.db

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Sora:wght@300;400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Sora', sans-serif; }

section[data-testid="stSidebar"] { background: #0f172a; border-right: 1px solid #1e293b; }
section[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
section[data-testid="stSidebar"] .stButton button {
    background: #1e293b; border: 1px solid #334155; color: #e2e8f0 !important;
    border-radius: 6px; font-size: 13px; text-align: left;
    width: 100%; margin-bottom: 4px; transition: all .15s;
}
section[data-testid="stSidebar"] .stButton button:hover { background: #334155; border-color: #6366f1; }
.main { background: #0f172a; }

.tracker-table { width: 100%; border-collapse: collapse; font-size: 13px; table-layout: auto; }
.tracker-table th {
    background: #1e293b; color: #94a3b8; padding: 10px 14px; border: 1px solid #334155;
    font-family: 'JetBrains Mono', monospace; font-size: 11px;
    text-transform: uppercase; letter-spacing: .08em; min-width: 160px; position: sticky; top: 0; z-index: 2;
}
.tracker-table td { border: 1px solid #1e293b; padding: 0; vertical-align: top; background: #0f172a; min-width: 160px; }
.tracker-table td.date-cell {
    background: #1e293b; color: #6366f1; font-family: 'JetBrains Mono', monospace;
    font-size: 12px; font-weight: 600; white-space: nowrap; padding: 10px 14px; min-width: 170px;
}
.tracker-table tr:hover td { background: #0d1525; }
.tracker-table tr:hover td.date-cell { background: #1a2744; }
.cell-view { padding: 8px 12px; min-height: 44px; cursor: pointer; transition: background .12s; }
.cell-view:hover { background: #1a2232 !important; }
.cell-view.active { background: #1e1b4b !important; outline: 2px solid #6366f1; outline-offset: -2px; }
.cell-placeholder { color: #334155; font-size: 12px; }

.tag-pill {
    display: inline-block; padding: 2px 9px; border-radius: 20px;
    font-size: 11px; font-weight: 600; margin: 2px 2px 2px 0; color: #fff; letter-spacing: .04em;
}
.save-banner {
    background: #1e1b4b; border: 1px solid #6366f1; border-radius: 8px;
    padding: 10px 16px; margin-bottom: 14px; color: #a5b4fc; font-size: 13px;
}
.stButton button {
    background: #1e293b !important; color: #e2e8f0 !important;
    border: 1px solid #334155 !important; border-radius: 6px !important;
}
.stButton button:hover { border-color: #6366f1 !important; background: #334155 !important; }
.stTextInput input, .stTextArea textarea {
    background: #1e293b !important; color: #e2e8f0 !important; border-color: #334155 !important;
}

/* Chat bubbles */
.chat-user {
    background: #1e1b4b; border: 1px solid #4338ca; border-radius: 12px 12px 2px 12px;
    padding: 12px 16px; margin: 8px 0 8px 15%; color: #e2e8f0; font-size: 14px; line-height: 1.6;
}
.chat-assistant {
    background: #0f2027; border: 1px solid #1e293b; border-radius: 12px 12px 12px 2px;
    padding: 12px 16px; margin: 8px 15% 8px 0; color: #cbd5e1; font-size: 14px; line-height: 1.6;
}
.chat-label-user { font-size: 11px; color: #6366f1; font-weight: 600; margin-bottom: 4px; text-align: right; margin-right: 4px; }
.chat-label-ai   { font-size: 11px; color: #22c55e; font-weight: 600; margin-bottom: 4px; margin-left: 4px; }
.chat-container  { max-height: 480px; overflow-y: auto; padding: 8px 0; }

.page-title { font-size: 26px; font-weight: 700; color: #e2e8f0; letter-spacing: -.02em; margin-bottom: 4px; }
.page-sub   { font-size: 12px; color: #475569; font-family: 'JetBrains Mono', monospace; margin-bottom: 16px; }
</style>
""", unsafe_allow_html=True)

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📊 Trade Tracker")
    st.markdown("---")

    with st.expander("➕ New Page", expanded=False):
        new_name = st.text_input("Page name", key="new_page_name", placeholder="e.g. NQ Contracts")
        if st.button("Create Page") and new_name.strip():
            name = new_name.strip()
            if name not in db:
                db[name] = {"columns": [], "rows": {}}
                save_data(db)
                st.session_state.current_page = name
                st.rerun()
            else:
                st.warning("Page already exists.")

    st.markdown("**Pages**")
    if not db:
        st.caption("No pages yet.")
    for pg in list(db.keys()):
        c1, c2 = st.columns([4, 1])
        with c1:
            if st.button(f"📄 {pg}", key=f"nav_{pg}"):
                st.session_state.current_page = pg
                st.session_state.active_cell = None
                st.session_state.edits = {}
                st.session_state.has_unsaved = False
                st.rerun()
        with c2:
            if st.button("🗑", key=f"del_{pg}"):
                del db[pg]
                if st.session_state.current_page == pg:
                    st.session_state.current_page = next(iter(db), None)
                save_data(db)
                st.rerun()

    st.markdown("---")
    st.markdown("**⬇️ Export Data**")
    if db:
        excel_data = export_to_excel(db)
        st.download_button(
            label="📊 Download as Excel",
            data=excel_data,
            file_name=f"trade_tracker_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    st.download_button(
        label="🗂 Download as JSON",
        data=json.dumps(db, indent=2, default=str),
        file_name=f"trade_tracker_{date.today()}.json",
        mime="application/json",
        use_container_width=True,
    )

    st.markdown("---")
    st.markdown("**Tag Legend**")
    for tag, color in TAGS.items():
        st.markdown(f'<span class="tag-pill" style="background:{color}">{tag}</span>', unsafe_allow_html=True)

# ── MAIN TABS ─────────────────────────────────────────────────────────────────
tab_tracker, tab_chat = st.tabs(["📊 Tracker", "🤖 AI Trade Analyst"])

# ════════════════════════════════════════════════════════════════════
# TAB 1 — TRACKER
# ════════════════════════════════════════════════════════════════════
with tab_tracker:
    page = st.session_state.current_page

    if page is None:
        st.markdown("""
        <div style="text-align:center;padding:80px 0;color:#475569;">
            <div style="font-size:52px;margin-bottom:16px;">📊</div>
            <div style="font-size:22px;font-weight:700;color:#94a3b8;margin-bottom:8px;">Trade Contract Tracker</div>
            <div style="font-size:14px;">Create a page from the sidebar to get started.</div>
        </div>""", unsafe_allow_html=True)
        st.stop()

    page_data = db[page]
    st.markdown(f'<div class="page-title">📄 {page}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="page-sub">observation tracker • {len(page_data["rows"])} rows</div>', unsafe_allow_html=True)

    # Universal Save
    if st.session_state.has_unsaved:
        sb1, sb2, _ = st.columns([2, 1, 6])
        with sb1:
            if st.button("💾 Save All Changes", type="primary"):
                for key, val in st.session_state.edits.items():
                    dstr, col = key.split("__", 1)
                    if dstr not in page_data["rows"]:
                        page_data["rows"][dstr] = {}
                    page_data["rows"][dstr][col] = val
                save_data(db)
                st.session_state.edits = {}
                st.session_state.has_unsaved = False
                st.session_state.active_cell = None
                st.success("✅ All changes saved!")
                st.rerun()
        with sb2:
            if st.button("✖ Discard"):
                st.session_state.edits = {}
                st.session_state.has_unsaved = False
                st.session_state.active_cell = None
                st.rerun()
        st.markdown(f'<div class="save-banner">⚠️ <b>{len(st.session_state.edits)}</b> unsaved cell(s). Click <b>Save All Changes</b> when done.</div>', unsafe_allow_html=True)

    # Manage Columns
    with st.expander("⚙️ Manage Columns", expanded=len(page_data["columns"]) == 0):
        cu1, cu2 = st.columns([4, 1])
        with cu1:
            new_col = st.text_input("Contract / column name", placeholder="e.g. NQ Dec-24", key="new_col")
        with cu2:
            st.write(""); st.write("")
            if st.button("Add Column") and new_col.strip():
                col_name = new_col.strip()
                if col_name not in page_data["columns"]:
                    page_data["columns"].append(col_name)
                    save_data(db)
                    st.rerun()
        if page_data["columns"]:
            remove_col = st.selectbox("Remove column", ["—"] + page_data["columns"], key="remove_col")
            if st.button("Remove") and remove_col != "—":
                page_data["columns"].remove(remove_col)
                for rd in page_data["rows"].values():
                    rd.pop(remove_col, None)
                save_data(db)
                st.rerun()

    # Add Date Row
    with st.expander("➕ Add Date Row", expanded=False):
        dc1, dc2 = st.columns([3, 1])
        with dc1:
            chosen_date = st.date_input("Pick a date", value=date.today(), key="date_picker")
        with dc2:
            st.write(""); st.write("")
            if st.button("Add Row"):
                dstr = str(chosen_date)
                if dstr not in page_data["rows"]:
                    page_data["rows"][dstr] = {}
                save_data(db)
                st.rerun()

    if not page_data["columns"]:
        st.info("Add at least one contract column above to start tracking.")
        st.stop()
    if not page_data["rows"]:
        st.info("Add a date row to begin entering observations.")
        st.stop()

    # Table
    sorted_dates = sorted(page_data["rows"].keys())
    active = st.session_state.active_cell
    header_html = '<th>📅 Date</th>' + "".join(f"<th>{c}</th>" for c in page_data["columns"])
    table_rows_html = ""

    for dstr in sorted_dates:
        row_data = page_data["rows"][dstr]
        try:
            d = datetime.strptime(dstr, "%Y-%m-%d")
            date_display = d.strftime("%a, %b %d %Y")
        except:
            date_display = dstr

        row_cells = f'<td class="date-cell">📅 {date_display}</td>'
        for col in page_data["columns"]:
            edit_key = f"{dstr}__{col}"
            if edit_key in st.session_state.edits:
                cell = st.session_state.edits[edit_key]
            else:
                cell = row_data.get(col, {"text": "", "tags": []})

            text = cell.get("text", "")
            tags = cell.get("tags", [])
            is_active = (active == (dstr, col))
            tags_html = "".join(f'<span class="tag-pill" style="background:{TAGS.get(t,"#555")}">{t}</span>' for t in tags)
            if tags or text:
                content_inner = f"{tags_html}<div style='margin-top:4px;color:#cbd5e1;font-size:13px;white-space:pre-wrap'>{text}</div>"
            else:
                content_inner = '<span class="cell-placeholder">click to edit</span>'
            active_class = "active" if is_active else ""
            row_cells += f'<td><div class="cell-view {active_class}">{content_inner}</div></td>'
        table_rows_html += f"<tr>{row_cells}</tr>"

    st.markdown(f"""
    <div style="overflow-x:auto;border-radius:10px;border:1px solid #1e293b;margin-bottom:4px;">
    <table class="tracker-table">
      <thead><tr>{header_html}</tr></thead>
      <tbody>{table_rows_html}</tbody>
    </table>
    </div>""", unsafe_allow_html=True)

    st.markdown("<div style='font-size:12px;color:#475569;margin-bottom:6px;'>👆 Select row & column to edit a cell</div>", unsafe_allow_html=True)
    sel_cols = st.columns([2, 2, 1])
    with sel_cols[0]:
        sel_date = st.selectbox("Row", sorted_dates, key="sel_date",
            format_func=lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%a, %b %d %Y"),
            label_visibility="collapsed")
    with sel_cols[1]:
        sel_col = st.selectbox("Column", page_data["columns"], key="sel_col", label_visibility="collapsed")
    with sel_cols[2]:
        if st.button("✏️ Open"):
            st.session_state.active_cell = (sel_date, sel_col)
            st.rerun()

    if active:
        adate, acol = active
        edit_key = f"{adate}__{acol}"
        if edit_key in st.session_state.edits:
            current = st.session_state.edits[edit_key]
        else:
            current = page_data["rows"].get(adate, {}).get(acol, {"text": "", "tags": []})
        try:
            cell_label = datetime.strptime(adate, "%Y-%m-%d").strftime("%a, %b %d %Y")
        except:
            cell_label = adate

        st.markdown(f"<div style='margin:12px 0 6px;font-size:13px;color:#a5b4fc;font-weight:600;'>✏️ Editing: <code>{cell_label}</code> × <b>{acol}</b></div>", unsafe_allow_html=True)
        ec1, ec2 = st.columns([3, 2])
        with ec1:
            new_text = st.text_area("Notes", value=current.get("text", ""), height=110,
                key=f"edit_text_{edit_key}", placeholder="Type your observation here…", label_visibility="collapsed")
        with ec2:
            new_tags = st.multiselect("Tags", options=list(TAGS.keys()),
                default=current.get("tags", []), key=f"edit_tags_{edit_key}", label_visibility="collapsed")
            if new_tags:
                st.markdown("".join(f'<span class="tag-pill" style="background:{TAGS[t]}">{t}</span>' for t in new_tags), unsafe_allow_html=True)

        draft = {"text": new_text, "tags": new_tags}
        if draft != current:
            st.session_state.edits[edit_key] = draft
            st.session_state.has_unsaved = True

        if st.button("Close ✖"):
            st.session_state.active_cell = None
            st.rerun()

    st.markdown("---")
    with st.expander("🗑️ Delete a Date Row"):
        del_date = st.selectbox("Select row", sorted_dates, key="del_date",
            format_func=lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%a, %b %d %Y"))
        if st.button("Delete Row", type="primary"):
            del page_data["rows"][del_date]
            save_data(db)
            st.rerun()

# ════════════════════════════════════════════════════════════════════
# TAB 2 — AI TRADE ANALYST
# ════════════════════════════════════════════════════════════════════
with tab_chat:
    st.markdown('<div class="page-title">🤖 AI Trade Analyst</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Ask Claude anything about your tracker data — it reads all your observations before answering.</div>', unsafe_allow_html=True)

    # Focus selector
    cf1, cf2 = st.columns([3, 1])
    with cf1:
        focus_options = ["All Pages"] + list(db.keys())
        focus = st.selectbox(
            "📂 Analyse data from:",
            focus_options,
            key="chat_focus_select",
            help="Narrow Claude's context to one page, or give it everything."
        )
    with cf2:
        st.write("")
        if st.button("🗑 Clear Chat"):
            st.session_state.chat_messages = []
            st.rerun()

    # Quick prompt buttons
    st.markdown("**Quick prompts:**")
    qp_cols = st.columns(4)
    quick_prompts = [
        "📈 Summarise all bullish signals",
        "📉 What bearish patterns do you see?",
        "🔁 Which contract has most Absorption tags?",
        "💡 Suggest a trade for tomorrow",
    ]
    for i, qp in enumerate(quick_prompts):
        with qp_cols[i]:
            if st.button(qp, key=f"qp_{i}", use_container_width=True):
                st.session_state.chat_messages.append({"role": "user", "content": qp})
                focus_page = None if focus == "All Pages" else focus
                context = build_tracker_context(db, focus_page)
                with st.spinner("Claude is analysing your data…"):
                    reply = ask_claude(st.session_state.chat_messages, context)
                st.session_state.chat_messages.append({"role": "assistant", "content": reply})
                st.rerun()

    st.markdown("---")

    # Chat history
    if not st.session_state.chat_messages:
        st.markdown("""
        <div style="text-align:center;padding:40px 0;color:#475569;">
            <div style="font-size:36px;margin-bottom:12px;">🤖</div>
            <div style="font-size:15px;color:#64748b;">Ask me anything about your trade observations.<br>
            I'll read your tracker data and give you specific, data-backed insights.</div>
        </div>""", unsafe_allow_html=True)
    else:
        chat_html = '<div class="chat-container">'
        for msg in st.session_state.chat_messages:
            if msg["role"] == "user":
                chat_html += f'<div class="chat-label-user">You</div><div class="chat-user">{msg["content"]}</div>'
            else:
                # convert newlines to <br> for HTML display
                content = msg["content"].replace("\n", "<br>")
                chat_html += f'<div class="chat-label-ai">🤖 Claude</div><div class="chat-assistant">{content}</div>'
        chat_html += '</div>'
        st.markdown(chat_html, unsafe_allow_html=True)

    # Input box
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    inp_col, btn_col = st.columns([6, 1])
    with inp_col:
        user_input = st.text_input(
            "Message",
            placeholder="e.g. What does the Iceberg activity in SRW suggest for next session?",
            key="chat_input",
            label_visibility="collapsed",
        )
    with btn_col:
        send = st.button("Send ➤", use_container_width=True)

    if send and user_input.strip():
        st.session_state.chat_messages.append({"role": "user", "content": user_input.strip()})
        focus_page = None if focus == "All Pages" else focus
        context = build_tracker_context(db, focus_page)
        with st.spinner("Claude is analysing your tracker data…"):
            reply = ask_claude(st.session_state.chat_messages, context)
        st.session_state.chat_messages.append({"role": "assistant", "content": reply})
        st.rerun()
